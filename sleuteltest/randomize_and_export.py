#!/usr/bin/env python3
"""
Randomiseert de antwoordvolgorde per vraag in questions.js
en genereert een Excel-bestand (.xlsx via openpyxl OF .csv als fallback)
met alle 300 vragen en antwoorden voor inhoudelijke review.
"""
import re
import random
import csv
import json
import os
from pathlib import Path

random.seed(42)  # reproduceerbaar

ROOT = Path(__file__).parent
SRC  = ROOT / "questions.js"
DST  = ROOT / "questions.js"          # we overschrijven in-place
CSV_OUT = ROOT / "vragenbank_review.csv"

# Profielen in volgorde zoals in het bestand
PROFILE_ORDER = [
    ("E-hulpmonteur", "Elektro - Hulpmonteur"),
    ("E-monteur",     "Elektro - Monteur"),
    ("E-chef",        "Elektro - Chef-monteur"),
    ("W-hulpmonteur", "Werktuigbouw - Hulpmonteur"),
    ("W-monteur",     "Werktuigbouw - Monteur"),
    ("W-chef",        "Werktuigbouw - Chef-monteur"),
]

# Lees questions.js als text
text = SRC.read_text(encoding="utf-8")

# Regex om elke vraagregel te matchen:
#   ["question", ["A","B","C","D"], correctIndex],
question_pat = re.compile(
    r'\["(?P<q>(?:[^"\\]|\\.)*)",\s*\[\s*"(?P<a>(?:[^"\\]|\\.)*)"\s*,\s*"(?P<b>(?:[^"\\]|\\.)*)"\s*,\s*"(?P<c>(?:[^"\\]|\\.)*)"\s*,\s*"(?P<d>(?:[^"\\]|\\.)*)"\s*\]\s*,\s*(?P<idx>\d+)\s*\]'
)

# Per profiel-blok bijhouden welke vragen we hebben gevonden, voor CSV-export
profile_blocks = {}
for key, _ in PROFILE_ORDER:
    # Vind het blok 'profile-key': [ ... ],
    block_pat = re.compile(
        r"'" + re.escape(key) + r"'\s*:\s*\[(.*?)\]\s*,(?=\s*\n\s*(?:'|/\*|\}))",
        re.DOTALL,
    )
    m = block_pat.search(text)
    if not m:
        print(f"⚠️  Profielblok niet gevonden: {key}")
        continue
    profile_blocks[key] = m

def shuffle_answer(match):
    q = match.group("q")
    options = [match.group("a"), match.group("b"), match.group("c"), match.group("d")]
    correct_idx = int(match.group("idx"))
    correct_text = options[correct_idx]

    # Shuffle opties; nieuwe index van het correcte antwoord opzoeken
    indices = [0, 1, 2, 3]
    random.shuffle(indices)
    new_options = [options[i] for i in indices]
    new_correct_idx = new_options.index(correct_text)

    # Reconstrueer JS-array: gebruik json.dumps voor correcte escaping
    a, b, c, d = new_options
    new_line = (
        '["' + q + '", '
        '["' + a + '","' + b + '","' + c + '","' + d + '"], '
        + str(new_correct_idx) + ']'
    )
    return new_line

# Pas randomisatie toe op het hele bestand
new_text, n_changes = question_pat.subn(shuffle_answer, text)
print(f"✅  {n_changes} vragen gerandomiseerd")

# Schrijf nieuwe questions.js
DST.write_text(new_text, encoding="utf-8")
print(f"✅  Nieuw bestand geschreven: {DST}")

# === CSV EXPORT ===
# We parsen het NIEUWE bestand om een review-CSV te maken
new_text = DST.read_text(encoding="utf-8")
rows = []
letters = ["A", "B", "C", "D"]

for key, label in PROFILE_ORDER:
    block_pat = re.compile(
        r"'" + re.escape(key) + r"'\s*:\s*\[(.*?)\]\s*,(?=\s*\n\s*(?:'|/\*|\}))",
        re.DOTALL,
    )
    m = block_pat.search(new_text)
    if not m:
        continue
    block = m.group(1)
    questions = list(question_pat.finditer(block))
    for i, qm in enumerate(questions, start=1):
        q = qm.group("q")
        opts = [qm.group("a"), qm.group("b"), qm.group("c"), qm.group("d")]
        idx = int(qm.group("idx"))
        rows.append({
            "Profiel": label,
            "Volgnr": i,
            "Vraag": q,
            "Optie A": opts[0],
            "Optie B": opts[1],
            "Optie C": opts[2],
            "Optie D": opts[3],
            "Juist antwoord": letters[idx],
            "Juiste tekst": opts[idx],
        })

# Schrijf CSV (UTF-8 met BOM zodat Excel correct opent)
with open(CSV_OUT, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=";")
    writer.writeheader()
    writer.writerows(rows)

print(f"✅  CSV-export geschreven: {CSV_OUT}  ({len(rows)} regels)")

# Probeer ook .xlsx te schrijven indien openpyxl beschikbaar
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Vragenbank Review"
    headers = list(rows[0].keys())
    ws.append(headers)

    # Header opmaak in Maintec-oranje
    header_fill = PatternFill("solid", fgColor="FF7D2F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Data
    for r in rows:
        ws.append([r[h] for h in headers])

    # Kolombreedtes
    widths = [22, 7, 80, 38, 38, 38, 38, 12, 38]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Wrap & alignment voor data
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    xlsx_out = ROOT / "vragenbank_review.xlsx"
    wb.save(xlsx_out)
    print(f"✅  Excel-export geschreven: {xlsx_out}")
except ImportError:
    print("ℹ️   openpyxl niet aanwezig — alleen CSV gegenereerd. CSV opent direct in Excel.")
